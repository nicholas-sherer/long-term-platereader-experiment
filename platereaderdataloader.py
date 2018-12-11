#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug 14 15:30:38 2018

@author: kuhlmanlab
"""

from collections import OrderedDict
import numpy as np
import openpyxl as pyxl


def iter_rows(worksheet):
    '''Return the rows of an excel sheet as lists'''
    for row in worksheet.iter_rows():
        yield [cell.value for cell in row]


def import_platereader(filepath, sample_layout, plate_shape=None):
    '''Load platereader spreadsheet into python.

    Turn times into a numpy array and OD data into a dictionary with the well
    names as keys and numpy arrays of OD data as floats.'''

    workbook = pyxl.load_workbook(filepath, read_only=True)
    worksheet = workbook[workbook.get_sheet_names()[0]]
    data = list(iter_rows(worksheet))

    times = np.array([float(value.strip('s'))/3600 for value in data[1]])
    i = 2
    wells = []
    while type(data[i][0]) is float:
        wells.append(np.array([value for value in data[i]]))
        i = i + 1

    if len(wells) != len(sample_layout):
        raise RuntimeError('''The number of well names in the plate layout
                           must be the same as the number of rows with OD
                           data in the excel spreadsheet.''')

    plate_shape = plate_shape
    if plate_shape is None:
        if len(sample_layout) == 12:
            plate_shape = (3, 4)
        elif len(sample_layout) == 24:
            plate_shape = (4, 6)
        elif len(sample_layout) == 48:
            plate_shape = (6, 8)
        elif len(sample_layout) == 96:
            plate_shape = (8, 12)
        else:
            raise RuntimeError('''The number of wells is not a standard 12,
                               24, 48, or 96 well plate. Please enter the shape
                               of the plate you used in (row, column) form as
                               a keyword argument "plate_shape".''')

    rows = plate_shape[0]
    columns = plate_shape[1]
    row_names = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    layout_order = [row_names[i]+str(j) for i in range(rows)
                    for j in range(1, columns+1)]

    wells_dict = OrderedDict()
    layout_dict = OrderedDict()
    for index, well in enumerate(sample_layout):
        wells_dict[well] = wells[index]
        layout_dict[well] = layout_order[index]

    plate_dict = {}
    plate_dict['times'] = times
    plate_dict['well_ODs'] = wells_dict
    plate_dict['layout'] = layout_dict
    return plate_dict
